# -*- coding: utf-8 -*-
"""Синхронизация to-do-list*.xlsx (Owner≈Аксенов) → Kanboard + привязка панелей To-Do."""
from __future__ import annotations

import hashlib
import json
import os
import re
import subprocess
import time
from calendar import timegm
from datetime import datetime
from pathlib import Path
from typing import Any

RUNTIME = Path(__file__).resolve().parents[1]
OWNER_NEEDLE = "аксенов"
PHP_SYNC = RUNTIME / "docker" / "kanboard_sync_todo_xlsx.php"
JSON_PAYLOAD = RUNTIME / "artifacts" / "demo" / "aksenov_todo_tasks.json"
STATE_FILE = RUNTIME / "artifacts" / "demo" / "todo-kanboard-sync-state.json"

# Панели дашборда ↔ файлы в папке Ту-ду
TODO_SITES: list[dict[str, str]] = [
    {
        "group_id": "berezniki-main",
        "site": "berezniki",
        "needle": "березник",
        "title": "Дела · Березники",
        "scope": "berezniki",
    },
    {
        "group_id": "satimola-main",
        "site": "satimola",
        "needle": "сатимол",
        "title": "Дела · Сатимола",
        "scope": "satimola",
    },
    {
        "group_id": "germany-main",
        "site": "germany",
        "needle": "герман",
        "title": "Дела · Германия",
        "scope": "germany",
    },
]


def archive_root() -> Path:
    return Path(os.environ.get("TMKI_REGULATIONS_ARCHIVE", r"D:\Курсор\СКРУ-2"))


def todo_folder() -> Path:
    root = archive_root()
    for name in ("Ту-ду", "Ту-Ду", "ту-ду", "To-do", "Todo"):
        p = root / name
        if p.is_dir():
            return p
    # fallback: любая папка с to-do-list*.xlsx
    for p in root.rglob("to-do-list*.xlsx"):
        if not p.name.startswith("~$"):
            return p.parent
    raise FileNotFoundError("Папка Ту-ду не найдена в архиве СКРУ-2")


def list_todo_xlsx() -> list[Path]:
    folder = todo_folder()
    hits = [
        p
        for p in folder.glob("to-do-list*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    ]
    if not hits:
        raise FileNotFoundError(f"to-do-list*.xlsx не найдены в {folder}")
    return sorted(hits, key=lambda p: p.name.lower())


def find_todo_xlsx() -> Path:
    """Совместимость: самый свежий to-do-list*.xlsx."""
    return max(list_todo_xlsx(), key=lambda p: p.stat().st_mtime)


def site_meta_for_path(path: Path) -> dict[str, str]:
    name = path.name.lower()
    for meta in TODO_SITES:
        if meta["needle"] in name:
            return meta
    stem = path.stem.lower().replace("to-do-list", "").strip(" -_")
    return {
        "group_id": f"todo-{_slug(stem) or 'misc'}",
        "site": _slug(stem) or "misc",
        "needle": stem,
        "title": f"Дела · {stem or path.stem}",
        "scope": "all",
    }


def _slug(text: str) -> str:
    s = re.sub(r"\s+", " ", (text or "").strip().lower())
    s = re.sub(r"[^\wа-яё\- ]+", "", s, flags=re.I)
    return s.replace(" ", "-")[:80] or "task"


def _status_ui(status: str) -> str:
    s = (status or "").strip().lower()
    mapping = {
        "planning": "Новая",
        "approved": "В работе",
        "pending review": "На согласовании",
        "in progress": "В работе",
        "completed": "Выполнена",
        "done": "Выполнена",
    }
    return mapping.get(s, status or "Новая")


def _priority_ui(priority: str) -> str:
    p = (priority or "").strip().lower()
    if p == "high":
        return "high"
    if p == "low":
        return "low"
    return "medium"


def _parse_workbook(path: Path, *, owner_only: bool = False) -> list[dict[str, Any]]:
    import openpyxl

    meta = site_meta_for_path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    headers: list[str] = []
    for i, row in enumerate(rows):
        vals = [str(c or "").strip().lower() for c in row]
        if "due date" in vals and "owner" in vals:
            header_idx = i
            headers = [str(c or "").strip() for c in row]
            break
    if header_idx is None:
        raise ValueError(f"Не найдены столбцы Due Date / Owner в {path.name}")

    out: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if not any(row):
            continue
        data = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        title = str(data.get("Project / Task") or data.get("Task") or "").strip()
        owner = str(data.get("Owner") or "").strip()
        if not title:
            continue
        if title.lower().startswith("task name"):
            continue
        if owner_only:
            if not owner or OWNER_NEEDLE not in owner.lower():
                continue
        elif not owner:
            continue

        due = data.get("Due Date")
        due_ts = 0
        due_iso = ""
        if isinstance(due, datetime):
            due_iso = due.strftime("%Y-%m-%d")
            due_ts = timegm((due.year, due.month, due.day, 12, 0, 0))
        elif due:
            due_iso = str(due)[:10]
            try:
                dt = datetime.fromisoformat(due_iso)
                due_ts = timegm((dt.year, dt.month, dt.day, 12, 0, 0))
            except Exception:
                due_ts = 0

        notes = str(data.get("Notes") or "").strip()
        status = str(data.get("Status") or "Planning").strip()
        priority = str(data.get("Priority") or "Normal").strip()
        due_human = ""
        if due_iso:
            try:
                y, m, d = due_iso.split("-")
                due_human = f"{int(d):02d}.{int(m):02d}.{y}"
            except Exception:
                due_human = due_iso
        sticky = notes or title
        site_label = meta.get("title", "").replace("Дела · ", "") or meta["site"]
        card_title = f"[{site_label}] {sticky}"
        if due_human:
            card_title = f"{card_title} · до {due_human}"
        desc = "\n".join(
            p
            for p in [
                f"Объект: {site_label}",
                f"Задача: {title}",
                f"Owner: {owner}",
                f"Due Date: {due_human or due_iso}" if (due_human or due_iso) else "",
                f"Status: {status}",
                f"Priority: {priority}",
                f"Notes: {notes}" if notes else "",
                f"Источник: {path.name}",
            ]
            if p
        )
        out.append(
            {
                "reference": f"xlsx-aksenov:{meta['site']}:{_slug(title)}",
                "base_title": title,
                "title": card_title,
                "owner": owner,
                "status": status,
                "priority": priority,
                "date_due": due_ts,
                "due_iso": due_iso,
                "description": desc,
                "notes": notes,
                "site": meta["site"],
                "group_id": meta["group_id"],
                "source_file": path.name,
                "source_path": str(path),
                "ui_title": sticky or title,
                "ui_status": _status_ui(status),
                "ui_priority": _priority_ui(priority),
            }
        )
    return out


def parse_aksenov_tasks(path: Path | None = None) -> list[dict[str, Any]]:
    if path is not None:
        return _parse_workbook(path, owner_only=True)
    tasks: list[dict[str, Any]] = []
    for p in list_todo_xlsx():
        try:
            tasks.extend(_parse_workbook(p, owner_only=True))
        except Exception:
            continue
    return tasks


def parse_site_tasks(path: Path) -> list[dict[str, Any]]:
    """Все строки с Owner (для превью панели)."""
    return _parse_workbook(path, owner_only=False)


def file_mtime(path: Path | None = None) -> float:
    if path is not None:
        return path.stat().st_mtime
    files = list_todo_xlsx()
    return max(p.stat().st_mtime for p in files)


def sources_mtime_map() -> dict[str, float]:
    return {str(p): p.stat().st_mtime for p in list_todo_xlsx()}


def sync_kanboard(tasks: list[dict[str, Any]] | None = None, *, path: Path | None = None) -> dict[str, Any]:
    files = [path] if path is not None else list_todo_xlsx()
    tasks = tasks if tasks is not None else parse_aksenov_tasks(path)
    JSON_PAYLOAD.parent.mkdir(parents=True, exist_ok=True)
    payload = [
        {
            "reference": t["reference"],
            "title": t["title"],
            "description": t["description"],
            "date_due": t["date_due"],
            "status": t["status"],
            "priority": t["priority"],
        }
        for t in tasks
    ]
    JSON_PAYLOAD.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    subprocess.run(
        ["docker", "cp", str(JSON_PAYLOAD), "tmki-kanboard:/tmp/aksenov_todo_tasks.json"],
        check=True,
        capture_output=True,
        text=True,
    )
    subprocess.run(
        ["docker", "cp", str(PHP_SYNC), "tmki-kanboard:/tmp/kanboard_sync_todo_xlsx.php"],
        check=True,
        capture_output=True,
        text=True,
    )
    proc = subprocess.run(
        ["docker", "exec", "tmki-kanboard", "php", "/tmp/kanboard_sync_todo_xlsx.php"],
        check=True,
        capture_output=True,
        text=True,
    )
    line = (proc.stdout or "").strip().splitlines()[-1] if proc.stdout else "{}"
    try:
        result = json.loads(line)
    except json.JSONDecodeError:
        result = {"ok": True, "raw": line}

    mtimes = {str(p): p.stat().st_mtime for p in files}
    state = {
        "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "sources": [str(p) for p in files],
        "source": str(files[0]) if files else "",
        "mtime": max(mtimes.values()) if mtimes else 0.0,
        "mtimes": mtimes,
        "tasks": [
            {
                "title": t["title"],
                "due": t["due_iso"],
                "owner": t["owner"],
                "reference": t["reference"],
                "source_file": t.get("source_file"),
            }
            for t in tasks
        ],
        "result": result,
    }
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {
        "ok": True,
        "sources": state["sources"],
        "source": state["source"],
        "mtime": state["mtime"],
        "tasks": state["tasks"],
        **result,
    }


def bind_todo_panels(
    todo_groups: list[dict[str, Any]] | None = None,
    todo_files: dict[str, Any] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Привязать панели To-Do к реальным xlsx из Ту-ду и обновить превью задач."""
    by_group = {m["group_id"]: m for m in TODO_SITES}
    groups = [dict(g) for g in (todo_groups or [])]
    files_out: dict[str, Any] = {k: list(v or []) for k, v in (todo_files or {}).items()}

    # гарантируем три панели
    existing_ids = {g.get("id") for g in groups}
    for meta in TODO_SITES:
        if meta["group_id"] not in existing_ids:
            groups.append(
                {
                    "id": meta["group_id"],
                    "title": meta["title"],
                    "scope": meta["scope"],
                    "items": [],
                }
            )

    path_by_group: dict[str, Path] = {}
    for path in list_todo_xlsx():
        meta = site_meta_for_path(path)
        path_by_group[meta["group_id"]] = path

    for group in groups:
        gid = str(group.get("id") or "")
        path = path_by_group.get(gid)
        if path is None:
            continue
        mtime = path.stat().st_mtime
        date_s = time.strftime("%Y-%m-%d", time.localtime(mtime))
        files_out[gid] = [
            {
                "date": date_s,
                "file": path.name,
                "text": f"Excel To-Do: {path.name}",
                "absolute_path": str(path),
                "relative_path": path.name,
                "source": "skru2-todo",
            }
        ]
        try:
            rows = parse_site_tasks(path)
        except Exception:
            rows = []
        if rows:
            group["items"] = [
                {
                    "title": r["ui_title"],
                    "status": r["ui_status"],
                    "priority": r["ui_priority"],
                    "owner": r["owner"],
                    "due": r["due_iso"],
                    "notes": r["notes"],
                    "task": r["base_title"],
                }
                for r in rows[:8]
            ]
        group["xlsx_path"] = str(path)
        group["xlsx_name"] = path.name
        meta = by_group.get(gid)
        if meta:
            group["title"] = meta["title"]
            group["scope"] = meta["scope"]

    return groups, files_out


def index_todo_file(path: Path | None = None) -> dict[str, Any]:
    """Опциональная переиндексация xlsx в skru-2."""
    import requests
    import psycopg2

    paths = [path] if path is not None else list_todo_xlsx()
    ollama = os.environ.get("OLLAMA_URL") or os.environ.get("OLLAMA_BASE_URL") or "http://localhost:11434"
    model = os.environ.get("OLLAMA_EMBEDDING_MODEL") or os.environ.get("TMKI_EMBEDDING_MODEL") or "nomic-embed-text"
    dim = int(os.environ.get("TMKI_EMBEDDING_DIMS") or os.environ.get("TMKI_EMBEDDING_DIM") or 768)
    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    cur = conn.cursor()
    total = 0
    docs: list[str] = []
    for xlsx in paths:
        import openpyxl

        wb = openpyxl.load_workbook(xlsx, data_only=True)
        ws = wb.active
        lines = [f"Файл: {xlsx.name}", "Таблица To-do list"]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                lines.append(" | ".join(cells))
        text = "\n".join(lines)
        chunks = [text[i : i + 1800].strip() for i in range(0, max(1, len(text)), 1800) if text[i : i + 1800].strip()]
        r = requests.post(
            f"{ollama.rstrip('/')}/api/embed",
            json={"model": model, "input": [c[:8000] for c in chunks], "truncate": True, "keep_alive": "24h"},
            timeout=300,
        )
        r.raise_for_status()
        vectors = r.json().get("embeddings") or []
        doc_id = "doc_" + hashlib.sha1(str(xlsx).encode("utf-8")).hexdigest()[:12]
        docs.append(doc_id)
        cur.execute("DELETE FROM chunks WHERE corpus_id='skru-2' AND doc_id=%s", (doc_id,))
        for idx, (content, emb) in enumerate(zip(chunks, vectors)):
            chunk_id = "skru2_todo_" + hashlib.sha1(f"{xlsx}|{idx}".encode()).hexdigest()[:24]
            emb_str = "[" + ",".join(str(float(x)) for x in emb) + "]"
            cur.execute(
                """
                INSERT INTO chunks (
                    chunk_id, corpus_id, doc_id, doc_path, content, embedding,
                    embedding_dim, page, section, has_table, metadata, indexed_at, updated_at
                ) VALUES (%s,'skru-2',%s,%s,%s,%s::vector,%s,%s,'to-do-list',true,%s::jsonb,NOW(),NOW())
                ON CONFLICT (chunk_id) DO UPDATE SET content=EXCLUDED.content, embedding=EXCLUDED.embedding, updated_at=NOW()
                """,
                (
                    chunk_id,
                    doc_id,
                    str(xlsx),
                    content,
                    emb_str,
                    dim,
                    idx + 1,
                    json.dumps({"source": xlsx.name}, ensure_ascii=False),
                ),
            )
            total += 1
    conn.commit()
    cur.close()
    conn.close()
    return {"indexed_chunks": total, "doc_ids": docs, "paths": [str(p) for p in paths]}


def sync_from_xlsx(*, reindex: bool = False) -> dict[str, Any]:
    files = list_todo_xlsx()
    tasks = parse_aksenov_tasks()
    result = sync_kanboard(tasks)
    if reindex:
        try:
            result["index"] = index_todo_file()
        except Exception as exc:  # noqa: BLE001
            result["index_error"] = str(exc)
    try:
        from tmki_demo.synthetic_docs import ensure_synthetic_tree, root_path

        ensure_synthetic_tree()
        copies = []
        for xlsx in files:
            meta = site_meta_for_path(xlsx)
            dest = root_path() / "To-do" / meta.get("title", meta["site"]).replace("Дела · ", "") / xlsx.name
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(xlsx.read_bytes())
            copies.append(str(dest))
        result["synthetic_copies"] = copies
    except Exception as exc:  # noqa: BLE001
        result["synthetic_error"] = str(exc)
    return result
